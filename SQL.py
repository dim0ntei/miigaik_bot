import os
import sqlite3
from openpyxl import load_workbook
from transliterate import translit


def table_name(excel):
    wb = load_workbook(excel)
    sheet = wb.active
    main_cell_value = sheet[1][0].value
    faculty = main_cell_value.split('факультета ')[1].split('на')[0]

    title = faculty.replace('-', '_').replace(' ', '_')
    title_v2 = title.replace('(', '_').replace(')', '_')
    en_title = translit(title_v2, language_code='ru', reversed=True).replace("'", '')
    # print(en_title)
    return en_title


def sem_number(excel):
    wb = load_workbook(excel)
    sheet = wb.active
    main_cell_value = sheet[1][0].value
    date = main_cell_value.split('на ')[1]
    if date == '10.09.2020':
        return '1'
    elif date == '01.02.2021':
        return '2'
    elif date == '10.09.2021':
        return '3'
    elif date == '01.02.2022':
        return '4'


def group_name(excel):
    wb = load_workbook(excel)
    sheet = wb.active
    main_cell_value = sheet[1][0].value
    group = main_cell_value.split('группы ')[1].split('факультета')[0].replace('\n', '')
    # print(group)
    return group


def direction(group):
    output_name = group.split('I')[0]
    return output_name


def first_name_splitter(name):
    first_name = name.split(' ')[1]
    return first_name


def last_name_splitter(name):
    last_name = name.split(' ')[0]
    return last_name


def fill_tables():
    conn = sqlite3.connect('list_of_students.db')
    cur = conn.cursor()

    os.chdir('directory_for_excel')
    all_excel_s = os.listdir()

    for excel_in_action in all_excel_s:
        ready_table_name = table_name(excel_in_action)

        cur.execute(
            f'CREATE TABLE IF NOT EXISTS {ready_table_name} (sem, direction, group_name, person_number, last_name, first_name, comment)'
        )

    for excel_in_action in all_excel_s:
        wb = load_workbook(excel_in_action)
        sheet = wb.active
        max_row = sheet.max_row + 1

        sem = sem_number(excel_in_action)
        group = group_name(excel_in_action)
        direct = direction(group)

        ready_table_name = table_name(excel_in_action)

        for row in range(3, max_row):
            first_name = first_name_splitter(sheet[row][1].value)
            last_name = last_name_splitter(sheet[row][1].value)
            m_dict = {'sem': sem,
                      'direction': direct,
                      'group_name': group,
                      'person_number': sheet[row][0].value,
                      'last_name': last_name,
                      'first_name': first_name}
            query_text = f"""insert into {ready_table_name} ({', '.join(m_dict.keys())}) VALUES ({', '.join([f"'{value}'" for value in m_dict.values()])}); """
            cur.execute(query_text)
    conn.commit()

    os.chdir('../')


fill_tables()
